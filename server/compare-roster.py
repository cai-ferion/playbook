#!/usr/bin/env python3
"""Compare BILLINGTEMPLATE.xlsx 04/04/26 OHR list against io_employees."""
import openpyxl
import requests
from datetime import datetime
from collections import defaultdict

# 1. Load BILLINGTEMPLATE.xlsx
print("Loading BILLINGTEMPLATE.xlsx...")
wb = openpyxl.load_workbook('/home/ubuntu/upload/BILLINGTEMPLATE.xlsx', data_only=True)
ws = wb['SRT_BILL']
print(f"Sheet: SRT_BILL | Rows: {ws.max_row} | Cols: {ws.max_column}")

# Header: date, ohr, srt_id, billing_name, srt_status, actual_vs_projection, role, planning_group
# Filter for date = 2026-04-04
billing_ohrs = {}  # ohr -> {name, role, pg, status}
for r in range(2, ws.max_row + 1):
    date_val = ws.cell(r, 1).value
    if date_val is None:
        continue
    # Parse date
    if isinstance(date_val, datetime):
        date_str = date_val.strftime('%Y-%m-%d')
    else:
        date_str = str(date_val)[:10]
    
    if date_str == '2026-04-04':
        ohr = str(ws.cell(r, 2).value or '').strip()
        name = str(ws.cell(r, 4).value or '').strip()
        status = str(ws.cell(r, 5).value or '').strip()
        actual_proj = str(ws.cell(r, 6).value or '').strip()
        role = str(ws.cell(r, 7).value or '').strip()
        pg = str(ws.cell(r, 8).value or '').strip()
        if ohr:
            billing_ohrs[ohr] = {
                'name': name, 'role': role, 'pg': pg,
                'status': status, 'actual_proj': actual_proj
            }

print(f"\nBILLINGTEMPLATE OHRs for 2026-04-04: {len(billing_ohrs)}")

# Show unique planning groups and roles
pgs = set(v['pg'] for v in billing_ohrs.values())
roles = set(v['role'] for v in billing_ohrs.values())
statuses = set(v['status'] for v in billing_ohrs.values())
print(f"Planning Groups: {sorted(pgs)}")
print(f"Roles: {sorted(roles)}")
print(f"Statuses: {sorted(statuses)}")

# 2. Load io_employees
print("\nLoading io_employees from API...")
resp = requests.get("http://localhost:3000/api/io/employees", params={'limit': '5000'})
employees = resp.json()
print(f"io_employees total: {len(employees)}")

emp_dict = {}  # ohr -> employee record
for e in employees:
    ohr = str(e.get('ohr_id', '')).strip()
    emp_dict[ohr] = e

active_emp_ohrs = set(ohr for ohr, e in emp_dict.items() 
                       if e.get('employement_status', '') == 'Active')
all_emp_ohrs = set(emp_dict.keys())

print(f"Active employees: {len(active_emp_ohrs)}")
print(f"All employees (incl inactive): {len(all_emp_ohrs)}")

billing_ohr_set = set(billing_ohrs.keys())

# 3. Compare
print("\n" + "="*80)
print("DISCREPANCY REPORT: BILLINGTEMPLATE (04/04/26) vs io_employees")
print("="*80)

# 3a. OHRs in billing but not in io_employees at all
billing_only = billing_ohr_set - all_emp_ohrs
print(f"\n--- OHRs in BILLINGTEMPLATE but NOT in io_employees: {len(billing_only)} ---")
for ohr in sorted(billing_only):
    b = billing_ohrs[ohr]
    print(f"  OHR: {ohr} | Name: {b['name']} | Role: {b['role']} | PG: {b['pg']} | Status: {b['status']}")

# 3b. Active io_employees NOT in billing template
active_not_in_billing = active_emp_ohrs - billing_ohr_set
print(f"\n--- Active io_employees NOT in BILLINGTEMPLATE: {len(active_not_in_billing)} ---")
for ohr in sorted(active_not_in_billing):
    e = emp_dict[ohr]
    print(f"  OHR: {ohr} | Name: {e.get('full_name','')} | Role: {e.get('actual_role','')} | PG: {e.get('planning_group','')} | Status: {e.get('employement_status','')}")

# 3c. OHRs in billing but marked inactive/exited in io_employees
billing_but_inactive = billing_ohr_set.intersection(all_emp_ohrs) - active_emp_ohrs
print(f"\n--- OHRs in BILLINGTEMPLATE but INACTIVE in io_employees: {len(billing_but_inactive)} ---")
for ohr in sorted(billing_but_inactive):
    b = billing_ohrs[ohr]
    e = emp_dict[ohr]
    print(f"  OHR: {ohr} | Name: {b['name']} | Billing Status: {b['status']} | DB Status: {e.get('employement_status','')}")

# 3d. Role mismatches
role_mismatches = []
for ohr in billing_ohr_set.intersection(all_emp_ohrs):
    b_role = billing_ohrs[ohr]['role']
    e_role = emp_dict[ohr].get('actual_role', '')
    if b_role != e_role:
        role_mismatches.append({
            'ohr': ohr,
            'name': billing_ohrs[ohr]['name'],
            'billing_role': b_role,
            'db_role': e_role
        })

print(f"\n--- Role Mismatches: {len(role_mismatches)} ---")
for m in sorted(role_mismatches, key=lambda x: x['ohr']):
    print(f"  OHR: {m['ohr']} | Name: {m['name']} | Billing: {m['billing_role']} | DB: {m['db_role']}")

# 3e. Planning Group mismatches (billing uses long-form PG, DB uses short-form)
# Map long-form to short-form for comparison
PG_MAP = {
    'MASA_MAFSA_CTR_SCALED_REVIEW': 'S-ABF',
    'MASA_MAFSA_CTR_CONTENT_SCORING': 'CS-ABF',
    'MASA_MAFSA_CTR_CSO': 'CSO_CTR',
    'MASA_MAFSA_CTR_FAD': 'FAD_CTR',
    'MASA_MAFSA_CTR_RECALL_MEASUREMENT': 'RECALL_MEASUREMENT_CTR',
    'MASA_MAFSA_CTR_SME': 'SME_CTR',
    'MASA_MAFSA_CTR_QPE': 'QPE_CTR',
}

pg_mismatches = []
for ohr in billing_ohr_set.intersection(all_emp_ohrs):
    b_pg = billing_ohrs[ohr]['pg']
    e_pg = emp_dict[ohr].get('planning_group', '')
    # Try mapping
    b_pg_short = PG_MAP.get(b_pg, b_pg)
    if b_pg_short != e_pg:
        pg_mismatches.append({
            'ohr': ohr,
            'name': billing_ohrs[ohr]['name'],
            'billing_pg': b_pg,
            'billing_pg_short': b_pg_short,
            'db_pg': e_pg
        })

print(f"\n--- Planning Group Mismatches: {len(pg_mismatches)} ---")
for m in sorted(pg_mismatches, key=lambda x: x['ohr']):
    print(f"  OHR: {m['ohr']} | Name: {m['name']} | Billing PG: {m['billing_pg']} -> {m['billing_pg_short']} | DB PG: {m['db_pg']}")

# Summary
print("\n" + "="*80)
print("SUMMARY")
print("="*80)
print(f"BILLINGTEMPLATE OHRs (04/04/26): {len(billing_ohrs)}")
print(f"io_employees (Active): {len(active_emp_ohrs)}")
print(f"io_employees (Total): {len(all_emp_ohrs)}")
print(f"In billing but NOT in io_employees: {len(billing_only)}")
print(f"Active io_employees NOT in billing: {len(active_not_in_billing)}")
print(f"In billing but INACTIVE in io_employees: {len(billing_but_inactive)}")
print(f"Role mismatches: {len(role_mismatches)}")
print(f"Planning Group mismatches: {len(pg_mismatches)}")
